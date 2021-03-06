# -*- coding:utf-8 -*-
"""
Written by Nina
"""
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string


def get_sheet(dir, sheet_name):
    """
    To get sheet
    :param dir: path
    :param sheet_name: sheet, eg, Sheet1
    :return:
    """
    workbook = load_workbook(dir)
    worksheet = workbook.get_sheet_by_name(sheet_name)  # sheet
    return worksheet


def get_suppliers(worksheet, column):
    """
    To get the lists of suppliers from the original excel
    :param column: the column in suppliers, eg, G
    :return: list of suppliers
    """
    supplier = []
    cols = worksheet[column]  # 供应商数据列表，类型为tuple
    for item in cols[1:]:
        if item.value and item.value not in supplier:
            supplier.append(item.value)
    print '-------------Success to get list of suppliers-----------------'
    return supplier


def set_header(worksheet, defaulting_font, defaulting_border, defaulting_alignment):
    """
    To set defaulting styles of the header
    """

    # 写表头并设置其样式
    for index_column in range(len(head)):
        worksheet.cell(row=1, column=index_column + 1, value=head[index_column])
        worksheet.cell(row=1, column=index_column + 1).fill = PatternFill(fill_type='solid', fgColor='92D050')
        worksheet.cell(row=1, column=index_column + 1).font = defaulting_font
        worksheet.cell(row=1, column=index_column + 1).border = defaulting_border
        worksheet.cell(row=1, column=index_column + 1).alignment = defaulting_alignment
        worksheet.column_dimensions[get_column_letter(index_column+1)].width = 20.0  # 1 -> 'A'

    # 设置行高
    worksheet.row_dimensions[1].height = 30.0
    # ------------------------------------------------------------------------------------
    # 设置不起作用，后续继续研究
    # worksheet.row_dimensions[1].fill = PatternFill(fill_type='solid', fgColor='92D050')
    # worksheet.row_dimensions[1].font = defaulting_font
    # worksheet.row_dimensions[1].border = defaulting_border
    # worksheet.row_dimensions[1].alignment = defaulting_alignment
    # ------------------------------------------------------------------------------------
    print '-------------Success to set header of the excel-----------------'

    return worksheet


def set_data(origin_worksheet, new_worksheet, supplier):
    """
    To write data into new_worksheet
    :return:
    """
    row_num = 1
    tag = []  # 记录合并单元格的行数
    for index, cell in enumerate(origin_worksheet.rows):
        if index:
            if cell[6].value == supplier:  # 获取到相匹配的供应商所在行
                cell[7].value = cell[7].value.strftime('%Y/%m/%d')  # 设置入库日日期格式
                cell[11].value = cell[11].value.strftime('%Y/%m/%d')  # 设置验收销售日日期格式
                if cell[18].value:  # 开票数量不为0时记录，后面用于合并单元格
                    tag.append(row_num+1)
                row_num += 1
                line = [col.value for col in cell]
                new_worksheet.append(line)

    for inx in range(len(tag)):
        pos_start = tag[inx]
        if inx == len(tag)-1:
            pos_end = new_worksheet.max_row
        else:
            pos_end = tag[inx + 1] - 1

        # 合并单元格，需要合并的单元格如下：
        # total quantity, usd amount, total carton quantity, gross weight, net weight, total volume
        merge_list = ['S', 'T', 'U', 'V', 'W', 'X']
        for num in merge_list:
            new_worksheet.merge_cells(start_row=pos_start, start_column=column_index_from_string(num),
                                      end_row=pos_end, end_column=column_index_from_string(num))
    return tag, new_worksheet


def set_styleofcell(worksheet, tag, cell_font, cell_border, cell_alignment):
    """
    To set styles of each cell
    """
    nums_index = 0
    for index, item in enumerate(worksheet.rows):  # 行循环 （0，(A1,B1,......）
        if index:
            if nums_index == len(tag):
                col = worksheet.max_row + 1
            else:
                col = tag[nums_index]
            if index == col - 1:
                nums_index += 1
            if nums_index % 2:
                for cell in item:  # 列循环，得到每个cell
                    cell.font = cell_font
                    cell.border = cell_border
                    cell.fill = PatternFill(fill_type='solid', fgColor='D0CECE')
                    cell.alignment = cell_alignment
            else:
                for cell in item:  # 列循环，得到每个cell
                    cell.font = cell_font
                    cell.border = cell_border
                    cell.fill = PatternFill(fill_type='solid', fgColor='D6E0E8')
                    cell.alignment = cell_alignment

if __name__ == "__main__":
    dir = r'xx\test.xlsx'
    sheet_name = 'Sheet1'
    column = 'G'
    sheet = get_sheet(dir, sheet_name)  # 原始数据表
    suppliers = get_suppliers(sheet, column)  # 供应商列表
    len_suppliers = len(suppliers)
    # 表头
    head = [u'订单号\n SLIP NUMBER',
            u'订单号\n P0 NUMBER',
            u'品番\n ITEM NO',
            u'出货数\n DELIVERY QUANTITY',
            u'箱数\n CARTON QUANTITY',
            u'批次号\n LOT NO',
            u'供应商\n SUPPLIER NAME',
            u'入库日\n STOCK IN DATE',
            u'系统运输方式\n DEFAULT SHIPMENT WAY',
            u'实际运输方式\n ACTUAL SHIPMENT WAY',
            u'目的地\n DESTINATION_CODE',
            u'检收销售日\n ETD DATE	',
            u'供应商代码\n SUPPLIER CODE',
            u'报关代码\n HS CODE',
            u'中文开票品名\n CHINESE NAME',
            u'开票单位\n UNIT',
            u'供应商发票号\n SUPPLIER INVOICE NO',
            u'IPO发票号\n INVOICE NO',
            u'开票数量\n TOTAL QUANTITY',
            u'USD AMOUNT',
            u'TOTAL CARTON QUANTITY',
            u'GROSS WEIGHT',
            u'NET WEIGHT',
            u'TOTAL VOLUME',
            u'预录单号',
            u'对账日期',
            u'开票金额',
            u'发票编号'
            ]
    # 设置字体样式
    font = Font(name=u'宋体', size=9)
    border = Border(left=Side(style='thin', color='FF000000'),
                    right=Side(style='thin', color='FF000000'),
                    top=Side(style='thin', color='FF000000'),
                    bottom=Side(style='thin', color='FF000000'))
    border_sep = Border(left=Side(style='thin', color='FF000000'),
                        right=Side(style='thin', color='FF000000'),
                        top=Side(style='medium', color='FF000000'),
                        bottom=Side(style='thin', color='FF000000'))
    alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
    for i in range(len_suppliers):
        wb_supplier = Workbook()
        sheet_supplier = wb_supplier.active
        dir_supplier = r'xxx\\' + suppliers[i] + '.xlsx'
        sheet_supplier = set_header(sheet_supplier, font, border, alignment)  # 设置表头
        (tag, sheet_supplier) = set_data(sheet, sheet_supplier, suppliers[i])  #
        set_styleofcell(sheet_supplier, tag, font, border, alignment)
        wb_supplier.save(dir_supplier)

    wb_supplier.close()
    

